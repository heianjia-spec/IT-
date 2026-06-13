"""Excel import/export logic for assets."""
from io import BytesIO
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse
from .models import Asset

# Fields available for export per asset type
EXPORT_FIELDS = {
    'common': [
        ('asset_number', '资产编号'), ('name', '资产名称'), ('asset_type', '资产类型'),
        ('category', '资产分类'), ('brand', '品牌'), ('spec_model', '规格型号'),
        ('serial_number', '序列号/SN码'), ('nc_number', 'NC编号'), ('status', '资产状态'),
        ('department', '使用部门'), ('responsible_person', '责任人'),
        ('room_name', '机房名称'), ('cabinet_no', '机柜编号'),
        ('storage_location', '存放位置描述'), ('purchase_date', '购置日期'),
        ('production_date', '生产日期'), ('purchase_amount', '购置金额'),
        ('depreciation_years', '折旧年限'), ('warranty_expiry', '保修截止日期'),
        ('warranty_status', '保修状态'), ('remarks', '备注'),
    ],
    'hardware': [
        ('cpu_model', 'CPU型号'), ('memory_info', '内存'), ('disk_size', '硬盘'),
        ('os', '操作系统'), ('has_monitor', '是否带显示器'), ('monitor_model', '显示器型号'),
        ('oob_address', '管理地址(OOB)'), ('nic_info', '网卡信息'),
        ('ip_address', 'IP地址'), ('mac_address', 'MAC地址'),
        ('port_count', '端口数量'), ('firmware', '固件版本'),
    ],
    'printer': [
        ('printer_type', '打印机类型'), ('is_color', '是否彩色'),
        ('is_duplex', '是否双面打印'), ('connect_type', '连接方式'),
        ('cartridge_model', '硒鼓/墨盒型号'),
    ],
    'network': [
        ('network_type', '网络设备类型'), ('port_type', '端口类型'),
        ('poe_support', '是否支持PoE'), ('poe_power', 'PoE功率(W)'),
        ('stackable', '是否可堆叠'), ('stack_id', '堆叠编号'),
        ('redundancy_power', '是否冗余电源'), ('rack_position', '机架位置'),
    ],
    'software': [
        ('software_type', '软件类型'), ('version', '版本号'),
        ('license_type', '授权类型'), ('license_key', '许可证号'),
        ('license_count', '授权数量'), ('used_count', '已用数量'),
        ('license_expiry', '许可到期日期'), ('platform', '运行平台'),
        ('installed_device', '安装设备'),
    ],
    'consumable': [
        ('consumable_type', '耗材类型'), ('unit', '单位'),
        ('current_stock', '当前库存'), ('min_stock', '最低库存预警'),
        ('applicable_device', '适用设备'), ('unit_price', '单价'),
        ('last_purchase_date', '上次采购日期'),
    ],
}


def generate_import_template(asset_type=None):
    """Generate an .xlsx template with headers for the given asset type."""
    wb = Workbook()
    ws = wb.active
    ws.title = '资产导入模板'

    headers = [h[1] for h in EXPORT_FIELDS['common']]
    if asset_type in ('server', 'pc'):
        headers += [h[1] for h in EXPORT_FIELDS['hardware']]
        headers += [h[1] for h in EXPORT_FIELDS['network'] if h[0] in ('rack_position', 'redundancy_power')]
    elif asset_type == 'printer':
        headers += [h[1] for h in EXPORT_FIELDS['hardware'] if h[0] in ('ip_address', 'mac_address', 'firmware', 'port_count')]
        headers += [h[1] for h in EXPORT_FIELDS['printer']]
    elif asset_type in ('network_device', 'switch'):
        headers += [h[1] for h in EXPORT_FIELDS['hardware'] if h[0] in ('ip_address', 'mac_address', 'firmware', 'port_count')]
        headers += [h[1] for h in EXPORT_FIELDS['network']]
    elif asset_type == 'software':
        headers += [h[1] for h in EXPORT_FIELDS['software']]
    elif asset_type == 'consumable':
        headers += [h[1] for h in EXPORT_FIELDS['consumable']]
    else:
        headers += [h[1] for h in EXPORT_FIELDS['hardware']]

    ws.append(headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_assets_from_excel(file, asset_type):
    """Parse .xlsx and return (success_list, error_list)."""
    wb = load_workbook(file)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return [], [{'row': 0, 'error': '文件中无数据'}]

    headers = [cell.value for cell in ws[1]]
    success = []
    errors = []

    for i, row in enumerate(rows, start=2):
        data = {}
        for j, header in enumerate(headers):
            if j < len(row) and header:
                data[header] = row[j] if row[j] is not None else ''

        try:
            asset = Asset.objects.create(
                name=str(data.get('资产名称', '')),
                asset_type=asset_type,
                asset_number='',  # auto-generated
                brand=str(data.get('品牌', '')),
                spec_model=str(data.get('规格型号', '')),
                serial_number=str(data.get('序列号/SN码', '')),
                nc_number=str(data.get('NC编号', '')),
                status=data.get('资产状态', 'in_stock'),
                responsible_person=str(data.get('责任人', '')),
                room_name=str(data.get('机房名称', '')),
                cabinet_no=str(data.get('机柜编号', '')),
                storage_location=str(data.get('存放位置描述', '')),
                remarks=str(data.get('备注', '')),
            )
            success.append(asset)
        except Exception as e:
            errors.append({'row': i, 'error': str(e)})

    return success, errors


def export_assets_to_excel(queryset, fields=None):
    """Export assets to .xlsx. fields is list of field names."""
    wb = Workbook()
    ws = wb.active
    ws.title = '资产导出'

    all_fields = EXPORT_FIELDS['common'] + EXPORT_FIELDS['hardware'] + \
                 EXPORT_FIELDS['printer'] + EXPORT_FIELDS['network'] + \
                 EXPORT_FIELDS['software'] + EXPORT_FIELDS['consumable']
    field_map = dict(all_fields)

    if fields:
        selected = [(f, field_map[f]) for f in fields if f in field_map]
    else:
        selected = all_fields

    ws.append([h[1] for h in selected])

    for asset in queryset:
        row_data = []
        for field_name, _ in selected:
            value = getattr(asset, field_name, '')
            if isinstance(value, (date, datetime)):
                value = value.isoformat()
            elif isinstance(value, bool):
                value = '是' if value else ('否' if value is False else '')
            elif value is None:
                value = ''
            row_data.append(value)
        ws.append(row_data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
